import openpyxl, xlrd, os, datetime

data_source = "data/data.xls"





class FormAlreadyExistsError(Exception):
    pass

class FormYearGroupMismatchError(Exception):
    pass

class IncorrectYearGroupError(Exception):
    pass


class FormGroup:


    def __init__(self, desc):
        self.__description = desc
        self.__infCount = 0
        self.__excCont = 0

    def add_event(self,event_line):

        print("Received event: {0}".format(event_line)
              )
        if "EXC" in event_line:
            self.__excCont += 1
        elif "INF" in event_line or "INC" in event_line:
            self.__infCount += 1

        print("Updated stats for form {0} - inf {1}; exc {2}".format(self.__description, self.__infCount, self.__excCont))

    def get_event_stats(self):
        return self.__infCount, self.__excCont

    def get_exc_count(self):
        return self.__excCont

    def get_inf_count(self):
        return self.__infCount

    def get_description(self):
        return self.__description

    def print_stats(self):
        print("Stats for form {0} - inf {1}; exc {2}".format(self.__description, self.__infCount, self.__excCont))

class YearGroup:

    def __init__(self, desc):
        self.__yg_exc_count = 0
        self.__yg_inf_count = 0
        self.__forms = {}
        self.__description = desc

    def add_form(self, f, key):

        print("Key I've been given: {0}, my yeargroup description: {1}".format(key, self.__description))

        if key[:2] != self.__description:
            raise FormYearGroupMismatchError

        if key not in self.__forms:
            self.__forms[key] = f
        else:
            raise FormAlreadyExistsError

    def refresh_yg_events_count(self):

        self.__yg_exc_count = 0
        self.__yg_inf_count = 0

        print("Refreshing stats for {}".format(self.__description))

        for f in self.__forms:
            print("Processing stats for {0}".format(self.__forms[f].get_description()))
            self.__yg_exc_count += self.__forms[f].get_exc_count()
            self.__yg_inf_count += self.__forms[f].get_inf_count()
            print("Yeargroup stats - inf: {0}, exc: {1}".format(self.__yg_inf_count, self.__yg_exc_count))

    def add_event(self, event_details):

        if event_details['year_code'] != self.__description:
            raise IncorrectYearGroupError

        form_code = event_details['form_code']

        self.__forms[form_code].add_event(event_details['event_code'])


    def __str__(self):
        return "Yeargroup: {0}; Contains {1} forms; Inf Count: {2}; Exc Count {3}".format(self.__description, len(self.__forms), self.__yg_inf_count, self.__yg_exc_count)

# Process the Excel spreadsheet, returning a list of all of the year groups that have been identified
def get_year_groups(d):

    year_groups = {}

    data = xlrd.open_workbook(d, on_demand=True).sheet_by_name("Sheet1")

    for row in range(1, data.nrows):
        form_code = data.cell(row, 4).value
        year_code = form_code[:2] # 5th column contains the form code in format YYH (Year House)


        # If the year group object doesn't exist yet, create it
        if year_code not in year_groups:
            year_groups[year_code] = YearGroup(year_code)

        # Assign the current year group to a variable for easier access
        current_yg = year_groups[year_code]

        # Attempt to add the relevant form to the year group - if it already exists an exception will be raised
        try:
            current_yg.add_form(FormGroup(form_code),form_code)
        except FormAlreadyExistsError:
            pass


    # year_groups['10'].add_form(FormGroup("TestForm"),'10')

    return year_groups

def process_events(d, ygs):

    data = xlrd.open_workbook(d, on_demand=True).sheet_by_name("Sheet1")

    for row in range(1, data.nrows):
        event_details = {'event_code': data.cell(row,0).value,
                         'form_code': data.cell(row, 4).value,
                         'year_code': data.cell(row, 4).value[:2]}


        ygs[event_details['year_code']].add_event(event_details)



year_groups = get_year_groups(data_source)

process_events(data_source, year_groups)

for y in year_groups:
    year_groups[y].refresh_yg_events_count()

for y in year_groups:
    print(year_groups[y])
