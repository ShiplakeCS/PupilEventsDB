import openpyxl, xlrd, os, datetime, sqlite3

data_source = "data/data.xls"
form_groups_data_source = "data/forms.xls"
db_file = "data/pupil_events.db"
db_schema = "data/setup.sql"


class DataHolder:

    def __init__(self):
        self.total_exc = 0
        self.total_inf = 0
        self.max_exc_staff = {'staff_code':None, 'count':0}
        self.max_inf_staff = {'staff_code':None, 'count':0}
        self.av_exc_per_staff = 0.0
        self.av_inf_per_staff = 0.0

class DatabaseHandler:

    def __init__(self, data_source, form_groups_data, db_schema, db_file):
        self.__data_source = data_source
        self.__db_cnx = sqlite3.connect(db_file)
        # Allow sqlite3 to return Row objects to queries (see http://flask.pocoo.org/docs/0.12/patterns/sqlite3/)
        self.__db_cnx.row_factory = sqlite3.Row
        self.__form_groups_data = form_groups_data
        self.__db_schema = db_schema
        self.__data = DataHolder()
        self.setup_db()

    def query_db(self, query, args=(), one=False):
        cur = self.__db_cnx.cursor()
        cur.execute(query, args)
        rv = cur.fetchall()
        cur.close()
        return (rv[0] if rv else None) if one else rv

    def setup_db(self):

        # Initialise with basic schema
        setup_sql = open(db_schema, 'r').read()
        self.__db_cnx.executescript(setup_sql)
        self.__db_cnx.commit()

        # Add formgroup data
        self.process_form_groups()

        # Add event data
        self.process_events()

    def get_year_groups_dictionary(self):

        return {'07':1, '08':2, '09':3, '10':4, '11':5, '12':6, '13':7}

    def get_house_dictionary(self):

        return {'B':1, 'E':2, 'O':3, 'S':4, 'W':5, 'L':6}

    def get_formgroup_dictionary(self):

        fg_ids = {}

        for form in self.query_db('select * from FormGroup'):
            fg_ids[form['Description']] = form['ID']

        return fg_ids

    def process_form_groups(self):

        # Add Form Groups data to FormGroup table

        form_groups = {}
        year_group_ids = self.get_year_groups_dictionary()
        house_ids = self.get_house_dictionary()

        cursor = self.__db_cnx.cursor()

        fg_data = xlrd.open_workbook(self.__form_groups_data, on_demand=True).sheet_by_name("Sheet1")

        for row in range(1, fg_data.nrows):
            # Check if form group exists in form_groups{}. If not, add it.

            form_code = fg_data.cell(row, 0).value

            if form_code not in form_groups:
                form_groups[form_code] = {'ID': 0, 'Size': 0}

            # Increase the size of the form group by 1
            form_groups[form_code]['Size'] += 1

        # Add each form group to the FormGroup table, update form_groups{} with the ID numbers for each FormGroup,
        # returned by the database
        for fg in form_groups:
            house_code = fg[2:]
            year_code = fg[:2]
            cursor.execute("INSERT INTO FormGroup (Description, Size, HouseID, YearGroupID) VALUES (?,?,?,?)",
                           (fg, form_groups[fg]['Size'], house_ids[house_code], year_group_ids[year_code]))
            form_groups[fg]['ID'] = cursor.lastrowid

        self.__db_cnx.commit()

    def process_events(self):

        cursor = self.__db_cnx.cursor()

        data = xlrd.open_workbook(self.__data_source, on_demand=True).sheet_by_name("Sheet1")

        added_staff_ids = {}
        added_formgroup_ids = {}
        house_ids = self.get_house_dictionary()
        yeargroup_ids = self.get_year_groups_dictionary()
        formgroup_ids = self.get_formgroup_dictionary()

        for row in range(1, data.nrows):
            # for row in range(1, 50):

            # shortcuts to each piece of data in the row

            event_code = data.cell(row, 0).value
            staff_code = data.cell(row, 2).value
            # date = data.cell(row, 3).value
            form_code = data.cell(row, 4).value
            year_code = form_code[:2]  # 5th column contains the form code in format YYH (Year House)
            house_code = form_code[2:]

            # check if the staff code exists in the staff dictionary, if not then it needs adding to the Staff table and
            # the staff dictionary

            if staff_code not in added_staff_ids:
                cursor.execute("INSERT INTO Staff (StaffCode) VALUES ('" + staff_code + "')")
                added_staff_ids[staff_code] = cursor.lastrowid

            # Process event data

            if "EXC" in event_code:
                event_type_id = 1
            else:
                event_type_id = 2

            cursor.execute("INSERT INTO Event (EventTypeID, StaffID, FormGroupID) VALUES (?,?,?)",
                           (event_type_id, added_staff_ids[staff_code], formgroup_ids[form_code]))

        self.__db_cnx.commit()

    def get_events_by_house_size(self):

        # For each house, find the number of excellence slips and divide by the size of the house

        print("House\t\tTotal EXC\t\tSize\t\tEXC pp")
        print("-"*60)

        for house in self.get_house_dictionary():

            size = self.query_db("SELECT sum(Size) as Size FROM `FormGroup` where HouseID = {}".format(self.get_house_dictionary()[house]), one=True)['Size']
            exc_count = self.query_db("select count(Event.ID) as exc_count from Event, FormGroup where Event.FormGroupID = FormGroup.ID and FormGroup.HouseID = {} and Event.EventTypeID = 1".format(self.get_house_dictionary()[house]), one=True)['exc_count']
            print("{0}\t\t\t{1}\t\t\t{2}\t\t\t{3:0.3f}".format(house, exc_count, size, exc_count/size))

    def refresh_summary_data(self):

        # Get total number of EXC events from database
        row = self.query_db('SELECT count(ID) as Num FROM Event WHERE EventTypeID = 1', one=True)
        self.__data.total_exc = row['Num']

        # Get total number of INF events from database
        row = self.query_db('SELECT count(ID) as Num FROM Event WHERE EventTypeID = 2', one=True)
        self.__data.total_inf = row['Num']

        # Get total number of staff and use to calculate averages
        row = self.query_db('SELECT count(ID) as Num FROM Staff', one=True)
        num_staff = row['Num']

        self.__data.av_exc_per_staff = self.__data.total_exc / num_staff
        self.__data.av_inf_per_staff = self.__data.total_inf / num_staff

        # Calculate average ratio
        self.__data.av_ratio_staff = self.__data.av_exc_per_staff / self.__data.av_inf_per_staff if self.__data.av_inf_per_staff != 0.00 else 0.00

        # Find staff with most excellence slips
        row = self.query_db('SELECT count(Event.ID) as Num, StaffID, StaffCode from Event,Staff where EventTypeID = 1 and Event.StaffID = Staff.ID group by StaffID ORDER BY count(Event.ID) DESC;', one=True)
        self.__data.max_exc_staff['staff_code'] = row['StaffCode']
        self.__data.max_exc_staff['count'] = row['Num']
        self.__data.max_exc_staff['ID'] = row['StaffID']

        # Find staff with most infractions
        row = self.query_db('SELECT count(Event.ID) as Num, StaffID, StaffCode from Event,Staff where EventTypeID = 2 and Event.StaffID = Staff.ID group by StaffID ORDER BY count(Event.ID) DESC;', one=True)
        self.__data.max_inf_staff['staff_code'] = row['StaffCode']
        self.__data.max_inf_staff['count'] = row['Num']
        self.__data.max_inf_staff['ID'] = row['StaffID']

    def close_db_cnx(self):
        self.__db_cnx.close()

dbh = DatabaseHandler(data_source, form_groups_data_source, db_schema, db_file)

dbh.refresh_summary_data()

dbh.get_events_by_house_size()

dbh.close_db_cnx()
