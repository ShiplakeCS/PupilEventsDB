from flask import Flask, render_template, request

import openpyxl, xlrd, os, datetime

app = Flask(__name__)

source_file = "data/data.xls"
updated = datetime.datetime.fromtimestamp(os.path.getmtime(source_file))

class Teacher:

    # __events = []
    __numINF = 0
    __numEXC = 0
    __numINC = 0
    __ratio = 0
    __subject = None

    def add_event(self, event):

        if "EXC" in event:
            self.__numEXC += 1

        elif "INF" in event:
            self.__numINF += 1

        elif "INC" in event:
            self.__numINC += 1

        self.refresh_ratio()

    def assign_subject(self):
        pass

    def refresh_ratio(self):
        try:
            self.__ratio = "%0.2f" % (self.__numEXC / self.__numINF)
        except ZeroDivisionError:
            self.__ratio = "No Infractions"

    def get_num_inf(self):
        return self.__numINF

    def get_num_inc(self):
        return self.__numINC

    def get_num_exc(self):
        return self.__numEXC

    def get_ratio(self):
        return self.__ratio

    def get_subject(self):
        pass

    def get_events(self):
         return {'Excellence slips':self.__numEXC,
                 'Infractions':self.__numINF,
                 'Incompletes':self.__numINC,
                 'Ratio':self.__ratio}



class EventsDB:

    __staff = {}
    __allstaffstats = []
    __totalInf = 0
    __totalExc = 0
    __totalInc = 0
    __totalRatio = 0
    __maxInf = 0
    __maxExc = 0
    __maxInfStaff = None
    __maxExcStaff = None

    def __init__(self):

        if source_file[-4:].lower() == ".csv":
            print("CSV file detected")
            self.load_CSV_data()

        elif source_file[-5:].lower() == ".xlsx":
            print("XLSX file detected")
            self.load_XLSX_data()

        elif source_file[-4:].lower() == ".xls":
            print("XLS file detected")
            self.load_XLS_data()

        else:
            print("No compatible source data file format found!")
            raise Exception

        self.refresh_summary()

    def load_CSV_data(self):

        events_file = open(source_file, "r")

        print("Processing CSV data found in {0}...".format(source_file))

        events_file.readline()

        for line in events_file:
            line = line.rstrip("\n")
            cols = line.split(",")

            if cols[2] in self.__staff.keys():
                pass
            else:
                self.__staff[cols[2]] = Teacher()

            self.__staff[cols[2]].add_event(cols[0])

    def load_XLSX_data(self):

        events_data = openpyxl.load_workbook(source_file, read_only=True).get_sheet_by_name("Sheet1")
        print("Processing XLSX data found in {0}...".format(source_file))

        for row in range(2, events_data.max_row + 1):
            # Read key info from the present row in the worksheet
            category_code = events_data['A' + str(row)].value
            staff_code = events_data['C' + str(row)].value

            # Update output so that users know what's going on
            print(row, ":", category_code,"-", staff_code)

            # Test whether an entry needs adding to the __staff dictionary for the present staff code
            if staff_code in self.__staff.keys():
                pass
            else:
                self.__staff[staff_code] = Teacher()

            # Add the event to the relevant Teacher in the __staff dictionary
            self.__staff[staff_code].add_event(category_code)

    def load_XLS_data(self):

        events_data = xlrd.open_workbook(source_file, on_demand=True).sheet_by_name("Sheet1")

        for row in range(events_data.nrows):

            # Read key info from the present row in the worksheet
            category_code = events_data.cell(row, 0).value
            staff_code = events_data.cell(row, 2).value

            # Test whether an entry needs adding to the __staff dictionary for the present staff code
            if staff_code in self.__staff.keys():
                pass
            else:
                self.__staff[staff_code] = Teacher()

            # Add the event to the relevant Teacher in the __staff dictionary
            self.__staff[staff_code].add_event(category_code)




    def refresh_ratio(self):
        try:
            self.__totalRatio = "%0.2f" % (self.__totalExc / self.__totalInf)
        except ZeroDivisionError:
            self.__totalRatio = "No Infractions"


    def refresh_summary(self):

        self.__totalInc = 0
        self.__totalExc = 0
        self.__totalInf = 0

        for s in self.__staff:
            self.__totalInf += self.__staff[s].get_num_inf()
            self.__totalExc += self.__staff[s].get_num_exc()
            self.__totalInc += self.__staff[s].get_num_inc()

        self.refresh_ratio()

        for s in self.__staff.keys():
            events = self.__staff[s].get_events()
            if events['Infractions'] > self.__maxInf:
                self.__maxInf = events['Infractions']
                self.__maxInfStaff = s
            if events['Excellence slips'] > self.__maxExc:
                self.__maxExc = events['Excellence slips']
                self.__maxExcStaff = s

        self.generate_all_staff_stats()

    def get_summary(self):
        return {"avEXC":"%0.2f" % (float(self.__totalExc) / len(self.__staff)),
                "avINF":"%0.2f" % (float(self.__totalInf) / len(self.__staff)),
                "avINC":"%0.2f" % (float(self.__totalInc) / len(self.__staff)),
                "maxINF":str(self.__maxInf),
                "maxINFStaff":self.__maxInfStaff,
                "maxEXC":str(self.__maxExc),
                "maxEXCStaff":self.__maxExcStaff,
                "ratio":self.__totalRatio,
                "n":len(self.__staff)}

    def get_staff_keys(self):

        return sorted(self.__staff.keys())

    def get_teacher(self, staffcode):
        return self.__staff[staffcode]
	
    def generate_all_staff_stats(self):
        print("Generating all staff stats...")

        for s in sorted(self.__staff.keys()):
            t = self.get_teacher(s)
            self.__allstaffstats.append({'code': s, 'inf': t.get_num_inf(), 'exc': t.get_num_exc()})
		

    def get_all_staff_stats(self):
        print("All staff stats requested")
        return self.__allstaffstats

program = EventsDB()

@app.route('/')
def showSummaryStats():
    return render_template("index.html", summary = program.get_summary(), updated = updated)

@app.route('/allstaff/')
def showAllStaffStats():
    return render_template("allstaff.html", stats = program.get_all_staff_stats())

@app.route('/<staffcode>/')
def showStaffDetails(staffcode):
    global updated
    try:
        teacher = program.get_teacher(staffcode.upper())
        return render_template("staffDetails.html", staffcode = staffcode, events = teacher.get_events(), summary = program.get_summary(), updated = updated)
    except KeyError:
        return "No teacher found with code " +  staffcode

@app.route('/<staff1>/<staff2>')
def compareStaff(staff1, staff2):
    t1events = program.get_teacher(staff1.upper()).get_events()
    t2events = program.get_teacher(staff2.upper()).get_events()
    # TODO: Improve the following by passing dictionaries for each member of staff, rather than 4 variables for each (i.e. 2 dictionries) - you will need to twaeak the HTML template to iterate over dictionary items in Flask's syntax
    return render_template("staffCompare.html", s1 = staff1, s2 = staff2, t1 = t1events, t2 = t2events)


app.run(debug=True, host='0.0.0.0', port=2000)
