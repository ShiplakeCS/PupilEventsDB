from flask import Flask, render_template, request, redirect, url_for

import openpyxl, xlrd, os, datetime

app = Flask(__name__)

data_source = "data/data.xls"
updated = datetime.datetime.fromtimestamp(os.path.getmtime(data_source))


class FileTypeError(Exception):
    pass


class Teacher:

    # __events = []
    __numINF = 0
    __numEXC = 0
    __numINC = 0
    __ratio = 0
    __subject = None

    def add_event(self, event):

        if "EXC" in event:
            self.__numEXC += 1

        elif "INF" in event:
            self.__numINF += 1

        elif "INC" in event:
            self.__numINC += 1

        self.refresh_ratio()

    def assign_subject(self):
        pass

    def refresh_ratio(self):
        try:
            self.__ratio = "%0.2f" % (self.__numEXC / self.__numINF)
        except ZeroDivisionError:
            self.__ratio = "No Infractions"

    def get_num_inf(self):
        return self.__numINF

    def get_num_inc(self):
        return self.__numINC

    def get_num_exc(self):
        return self.__numEXC

    def get_ratio(self):
        return self.__ratio

    def get_subject(self):
        pass

    def get_events(self):
         return {'Excellence slips':self.__numEXC,
                 'Infractions':self.__numINF,
                 'Incompletes':self.__numINC,
                 'Ratio':self.__ratio}

class YearGroup:
	__id = ""
	__forms = {}
	__totalExc = 0
	__totalInf = 0
	__totalRatio = 0
	
	def __init__(self, y):
		self.__id = y # yeargroup, e.g. 7, 8, 9, etc
		
	def add_form(self, form_code):
		if form_code in self.__forms:
			return self.__forms[form_code]
		else:
			self.__forms[form_code] = Form(form_code)
			return self.__forms[form_code]
			
	def get_form(self, form_code):
		return self.__forms[form_code]
	
	def get_forms(self, ):
		return self.__forms
	
	def add_form_event(self, form_code, event):
		add_form(form_code).add_event(event)
	
	def update_event_data(self, event):
		if "EXC" in event:
			self.__totalExc += 1
		elif "INF" in event:
			self.__totalInf += 1
		elif "INC" in event:
			self.__totalInf += 1

        try:
            self.__totalRatio = "%0.2f" % (self.__totalExc / self.__totalInf)

        except ZeroDivisionError:
            self.__totalRatio = "No Infractions"
		
class Form:
	__id = ""
	__exc = 0
	__inf = 0
	__ratio = "0.0"
	
	def __init__(self, code):
		self.__id = code
		
	def add_event(self, event):

        if "EXC" in event:
            self.__exc += 1

        elif "INF" in event:
            self.__inf += 1

        elif "INC" in event:
            self.__inf += 1
		
		self.refresh_ratio(self)
		
	def refresh_ratio(self):
        try:
            self.__ratio = "%0.2f" % (self.__exc / self.__inf)
        except ZeroDivisionError:
            self.__ratio = "No Infractions"
	
	def get_events(self):
		"""Returns a dictionary of events (keys: exc, inf)"""
		return {'exc':self.__exc, 'inf':self.__inf}
	

class EventsDB:

    __staff = {}
    __allstaffstats = []
    __totalInf = 0
    __totalExc = 0
    __totalInc = 0
    __totalRatio = 0
    __maxInf = 0
    __maxExc = 0
    __maxInfStaff = None
    __maxExcStaff = None
    # Need to add year group dictionary

    def __init__(self):

        self.load_data(data_source)

    def clear_data(self):
        self.__staff = {}
        self.__allstaffstats = []
        self.__totalExc = 0
        self.__totalInc = 0
        self.__totalInf = 0
        self.__totalRatio = 0
        self.__maxExc = 0
        self.__maxInf = 0
        self.__maxExcStaff = None
        self.__maxInfStaff = None

    def load_data(self, source_file):

        self.clear_data()

        try:
            print("Loading source data...")
        except:
            pass

        if source_file[-4:].lower() == ".csv":
            try:
                print("CSV file detected")
            except:
                pass
            self.load_CSV_data(source_file)

        elif source_file[-5:].lower() == ".xlsx":
            try:
                print("XLSX file detected")
            except:
                pass
            self.load_XLSX_data(source_file)

        elif source_file[-4:].lower() == ".xls":
            try:
                print("XLS file detected")
            except:
                pass
            self.load_XLS_data(source_file)

        else:
            try:
                print("No compatible source data file format found!")
            except:
                pass
            raise Exception

        self.refresh_summary()

    def load_CSV_data(self, source_file):

        global updated

        events_file = open(source_file, "r")

        try:
            print("Processing CSV data found in {0}...".format(source_file))
        except:
            pass

        events_file.readline()

        for line in events_file:
            line = line.rstrip("\n")
            cols = line.split(",")

            if cols[2] in self.__staff.keys():
                pass
            else:
                self.__staff[cols[2]] = Teacher()

            self.__staff[cols[2]].add_event(cols[0])

			form_code = cols[4]

        updated = datetime.datetime.fromtimestamp(os.path.getmtime(source_file))

    def load_XLSX_data(self, source_file):

        global updated

        events_data = openpyxl.load_workbook(source_file, read_only=True).get_sheet_by_name("Sheet1")
        try:
            print("Processing XLSX data found in {0}...".format(source_file))
        except:
            pass

        for row in range(2, events_data.max_row + 1):
            # Read key info from the present row in the worksheet
            category_code = events_data['A' + str(row)].value
            staff_code = events_data['C' + str(row)].value
			form_code = events_data['E' + str(row)].value
			
            # Update output so that users know what's going on
            try:
                print(row, ":", category_code,"-", staff_code)
            except:
                pass

            # Test whether an entry needs adding to the __staff dictionary for the present staff code
            if staff_code in self.__staff.keys():
                pass
            else:
                self.__staff[staff_code] = Teacher()

            # Add the event to the relevant Teacher in the __staff dictionary
            self.__staff[staff_code].add_event(category_code)

        updated = datetime.datetime.fromtimestamp(os.path.getmtime(source_file))

    def load_XLS_data(self, source_file):

        global updated

        events_data = xlrd.open_workbook(source_file, on_demand=True).sheet_by_name("Sheet1")

        for row in range(events_data.nrows):

            # Read key info from the present row in the worksheet
            category_code = events_data.cell(row, 0).value
            staff_code = events_data.cell(row, 2).value
            form_code = events_data.cell(row, 4).value

            # Test whether an entry needs adding to the __staff dictionary for the present staff code
            if staff_code in self.__staff.keys():
                pass
            else:
                self.__staff[staff_code] = Teacher()

            # Add the event to the relevant Teacher in the __staff dictionary
            self.__staff[staff_code].add_event(category_code)

        updated = datetime.datetime.fromtimestamp(os.path.getmtime(source_file))

    def refresh_ratio(self):
        try:
            self.__totalRatio = "%0.2f" % (self.__totalExc / self.__totalInf)
        except ZeroDivisionError:
            self.__totalRatio = "No Infractions"

    def refresh_summary(self):

        self.__totalInc = 0
        self.__totalExc = 0
        self.__totalInf = 0

        for s in self.__staff:
            self.__totalInf += self.__staff[s].get_num_inf()
            self.__totalExc += self.__staff[s].get_num_exc()
            self.__totalInc += self.__staff[s].get_num_inc()

        self.refresh_ratio()

        for s in self.__staff.keys():
            events = self.__staff[s].get_events()
            if events['Infractions'] > self.__maxInf:
                self.__maxInf = events['Infractions']
                self.__maxInfStaff = s
            if events['Excellence slips'] > self.__maxExc:
                self.__maxExc = events['Excellence slips']
                self.__maxExcStaff = s

        self.generate_all_staff_stats()

    def get_summary(self):
        return {"avEXC":"%0.2f" % (float(self.__totalExc) / len(self.__staff)),
                "avINF":"%0.2f" % (float(self.__totalInf) / len(self.__staff)),
                "avINC":"%0.2f" % (float(self.__totalInc) / len(self.__staff)),
                "maxINF":str(self.__maxInf),
                "maxINFStaff":self.__maxInfStaff,
                "maxEXC":str(self.__maxExc),
                "maxEXCStaff":self.__maxExcStaff,
                "ratio":self.__totalRatio,
                "n":len(self.__staff)}

    def get_staff_keys(self):

        return sorted(self.__staff.keys())

    def get_teacher(self, staffcode):
        return self.__staff[staffcode]

    def generate_all_staff_stats(self):
        try:
            print("Generating all staff stats...")
        except:
            pass

        for s in sorted(self.__staff.keys()):
            t = self.get_teacher(s)
            self.__allstaffstats.append({'code': s, 'inf': t.get_num_inf(), 'exc': t.get_num_exc()})

    def get_all_staff_stats(self):
        try:
            print("All staff stats requested")
        except:
            pass

        return self.__allstaffstats

program = EventsDB()


@app.route('/')
def showSummaryStats():
    return render_template("index.html", summary = program.get_summary(), updated = updated)


@app.route('/allstaff/')
def showAllStaffStats():
    return render_template("allstaff.html", stats = program.get_all_staff_stats(), updated = updated)


@app.route('/<staffcode>/')
def showStaffDetails(staffcode):
    global updated
    try:
        teacher = program.get_teacher(staffcode.upper())
        return render_template("staffDetails.html", staffcode = staffcode, events = teacher.get_events(), summary = program.get_summary(), updated = updated)
    except KeyError:
        return "No teacher found with code " +  staffcode


@app.route('/<staff1>/<staff2>')
def compareStaff(staff1, staff2):
    t1events = program.get_teacher(staff1.upper()).get_events()
    t2events = program.get_teacher(staff2.upper()).get_events()
    # TODO: Improve the following by passing dictionaries for each member of staff, rather than 4 variables for each (i.e. 2 dictionries) - you will need to twaeak the HTML template to iterate over dictionary items in Flask's syntax
    return render_template("staffCompare.html", s1 = staff1, s2 = staff2, t1 = t1events, t2 = t2events)


@app.route('/update-data')
def upload_file():
    return render_template('update-data.html')


@app.route('/uploader', methods=['GET', 'POST'])
def update_data():
    if request.method == 'POST':
        f = request.files['file']
        if f.filename[-4:] == ".csv":
            filepath = "data/data.csv"

        elif f.filename[-4:] == ".xls":
            filepath = "data/data.xls"

        elif f.filename[-5:] == ".xlsx":
            filepath = "data/data.xlsx"
        else:
            raise FileTypeError("Filetype not recognised. Must use CSV, XLS or XLSX.")

        f.save(filepath)
        program.load_data(filepath)

        return redirect('/')

app.run(debug=True, host='0.0.0.0', port=2000)
