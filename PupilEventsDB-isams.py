from flask import Flask, render_template, request, redirect, url_for

import openpyxl, xlrd, os, datetime

app = Flask(__name__)

data_source = "data/report.xls"
updated = datetime.datetime.fromtimestamp(os.path.getmtime(data_source))


class FileTypeError(Exception):
    pass


# TODO: Add Form class, must be able to report numbers of inf and exc
# TODO: Add Year class that contains form objects
# TODO: Add House class that contains form objects

class FormGroup:

    def __init__(self):
        self.__inf = 0
        self.__exc = 0

    def addEvent(self, event):
        if "Excellence" in event:
            self.__exc += 1
        elif "Infraction" in event or "INC" in event:
            self.__inf += 1

    def getExc(self):
        return self.__exc

    def getInf(self):
        return self.__inf

    def getEventsCount(self):
        """Returns a tuple of exc, inf"""
        return self.__exc, self.__inf

class YearGroup:


    def __init__(self, yg):
        self.__yearGroup = yg
        self.__inf = 0
        self.__exc = 0
        self.__forms = {}


    def getForms(self):
        return self.__forms.keys()

    def addEvent(self, event, form_code):
        if form_code[:2] != self.__yearGroup:
            print("Not for this year")
            return


        if form_code not in self.__forms:
            print("No form found in this year")
        else:
            self.__forms[form_code].addEvent(event)
            #print("added event to form", form_code, "in yeargroup", self.__yearGroup)


    def addForm(self, f):
        if f[:2] != self.__yearGroup:
            print("Form {0} does not belong to year {1}!".format(f, self.year))
        else:
            self.__forms[f] = FormGroup()
            print("Form {0} created in Yeargroup {1}".format(f, self.__yearGroup))

    def printDetails(self):

        print("Yeargroup", self.__yearGroup)

        for f in self.__forms:
            print("Form {0}, inf: {1}, exc: {2}".format(f, self.__forms[f].getInf(), self.__forms[f].getExc()))

        print("Yeargroup {0} - Inf: {1}, Exc: {2}".format(self.__yearGroup, self.__inf, self.__exc))



    def refreshValues(self):

        self.__inf = 0
        self.__exc = 0

        for f in self.__forms:
            self.__exc = self.__exc +  self.__forms[f].getExc()
            self.__inf = self.__inf + self.__forms[f].getInf()

        print("Refreshed values for yeargroup {0} / inf: {1}, exc: {2}".format(self.__yearGroup, self.__inf, self.__exc))



class Teacher:

    def __init__(self):
        self.__numINF = 0
        self.__numEXC = 0
        self.__numINC = 0
        self.__ratio = 0
        self.__subject = None

    def add_event(self, event):

        if "Excellence" in event:
            self.__numEXC += 1

        elif "Infraction" in event:
            self.__numINF += 1

        elif "INC" in event:
            self.__numINC += 1

        self.refresh_ratio()

    def assign_subject(self):
        pass

    def refresh_ratio(self):
        try:
            self.__ratio = "%0.2f" % (self.__numEXC / self.__numINF)
        except ZeroDivisionError:
            self.__ratio = "No Infractions"

    def get_num_inf(self):
        return self.__numINF

    def get_num_inc(self):
        return self.__numINC

    def get_num_exc(self):
        return self.__numEXC

    def get_ratio(self):
        return self.__ratio

    def get_subject(self):
        pass

    def get_events(self):
         return {'Excellence slips':self.__numEXC,
                 'Infractions':self.__numINF,
                 'Incompletes':self.__numINC,
                 'Ratio':self.__ratio}


class EventsDB:

    ## Changes for yeargroup and form group calculations



    def __init__(self):

        self.__staff = {}
        self.__yearGroups = {}
        self.__allstaffstats = []
        self.__totalExc = 0
        self.__totalInc = 0
        self.__totalInf = 0
        self.__totalRatio = 0
        self.__maxExc = 0
        self.__maxInf = 0
        self.__maxExcStaff = None
        self.__maxInfStaff = None


    def clear_data(self):
        self.__staff = {}
        self.__yearGroups = {}
        self.__allstaffstats = []
        self.__totalExc = 0
        self.__totalInc = 0
        self.__totalInf = 0
        self.__totalRatio = 0
        self.__maxExc = 0
        self.__maxInf = 0
        self.__maxExcStaff = None
        self.__maxInfStaff = None

    def load_data(self, source_file):

        self.clear_data()

        try:
            print("Loading source data...")
        except:
            pass

        if source_file[-4:].lower() == ".xls":

            self.load_XLS_data(source_file)

        else:
            try:
                print("No compatible source data file format found!")
            except:
                pass
            raise Exception

    def load_XLS_data(self, source_file):

        global updated

        #events_data = xlrd.open_workbook(source_file, on_demand=True).sheet_by_name("Sheet1")
        events_data = xlrd.open_workbook(source_file, on_demand=True).sheet_by_name("Rewards Report")

        for row in range(1, events_data.nrows):

            # Read key info from the present row in the worksheet
            category_code = events_data.cell(row, 5).value
            staff_code = events_data.cell(row, 10).value
            form_code = events_data.cell(row, 3).value
            year_code = form_code[:2]
            house_code = form_code[-1:]

            # Test whether row contains meaningful data
            if "Excellence" in category_code or "Infraction" in category_code:

                # Test whether an entry needs adding to the __staff dictionary for the present staff code
                if staff_code in self.__staff.keys():
                    pass
                else:
                    self.__staff[staff_code] = Teacher()

                # Add the event to the relevant Teacher in the __staff dictionary
                self.__staff[staff_code].add_event(category_code)

                # Test if a relevant year group exists and, if not, create it

                if year_code not in self.__yearGroups.keys():
                    self.__yearGroups[year_code] = YearGroup(year_code)
                    #print("Year group created for {0}".format(year_code))

                # Test if the relevant year group contains the form
                if form_code not in self.__yearGroups[year_code].getForms() and form_code[:2] == year_code:
                    self.__yearGroups[year_code].addForm(form_code)
                    #print("Form group created for {0} in year {1}".format(form_code, year_code))

                # Add event to current yeargroup and form

                self.__yearGroups[year_code].addEvent(category_code, form_code)
                #print("Added {2} event to form {0} in year {1}".format(form_code, year_code, category_code))

        for yg in self.__yearGroups:
            self.__yearGroups[yg].refreshValues()
            #self.__yearGroups[yg].printDetails()

        updated = datetime.datetime.fromtimestamp(os.path.getmtime(source_file))

    def refresh_ratio(self):
        try:
            self.__totalRatio = "%0.2f" % (self.__totalExc / self.__totalInf)
        except ZeroDivisionError:
            self.__totalRatio = "No Infractions"

    def refresh_summary(self):

        self.__totalInc = 0
        self.__totalExc = 0
        self.__totalInf = 0

        for s in self.__staff:
            self.__totalInf += self.__staff[s].get_num_inf()
            self.__totalExc += self.__staff[s].get_num_exc()
            self.__totalInc += self.__staff[s].get_num_inc()

        self.refresh_ratio()

        for s in self.__staff.keys():
            events = self.__staff[s].get_events()
            if events['Infractions'] > self.__maxInf:
                self.__maxInf = events['Infractions']
                self.__maxInfStaff = s
            if events['Excellence slips'] > self.__maxExc:
                self.__maxExc = events['Excellence slips']
                self.__maxExcStaff = s

        self.generate_all_staff_stats()

    def get_summary(self):
        return {"avEXC":"%0.2f" % (float(self.__totalExc) / len(self.__staff)),
                "avINF":"%0.2f" % (float(self.__totalInf) / len(self.__staff)),
                "avINC":"%0.2f" % (float(self.__totalInc) / len(self.__staff)),
                "maxINF":str(self.__maxInf),
                "maxINFStaff":self.__maxInfStaff,
                "maxEXC":str(self.__maxExc),
                "maxEXCStaff":self.__maxExcStaff,
                "ratio":self.__totalRatio,
                "n":len(self.__staff)}

    def get_staff_keys(self):

        return sorted(self.__staff.keys())

    def get_teacher(self, staffcode):
        return self.__staff[staffcode]

    def generate_all_staff_stats(self):
        try:
            print("Generating all staff stats...")
        except:
            pass

        for s in sorted(self.__staff.keys()):
            t = self.get_teacher(s)
            self.__allstaffstats.append({'code': s, 'inf': t.get_num_inf(), 'exc': t.get_num_exc()})

    def get_all_staff_stats(self):
        try:
            print("All staff stats requested")
        except:
            pass

        return self.__allstaffstats


#program.load_data(data_source)


@app.route('/')
def showSummaryStats():
    return render_template("index.html", summary = program.get_summary(), updated = updated)


@app.route('/allstaff/')
def showAllStaffStats():
    return render_template("allstaff.html", stats = program.get_all_staff_stats(), updated = updated)


@app.route('/<staffcode>/')
def showStaffDetails(staffcode):
    global updated
    try:
        teacher = program.get_teacher(staffcode.upper())
        return render_template("staffDetails.html", staffcode = staffcode, events = teacher.get_events(), summary = program.get_summary(), updated = updated)
    except KeyError:
        return "No teacher found with code " +  staffcode


@app.route('/<staff1>/<staff2>')
def compareStaff(staff1, staff2):
    t1events = program.get_teacher(staff1.upper()).get_events()
    t2events = program.get_teacher(staff2.upper()).get_events()
    # TODO: Improve the following by passing dictionaries for each member of staff, rather than 4 variables for each (i.e. 2 dictionries) - you will need to twaeak the HTML template to iterate over dictionary items in Flask's syntax
    return render_template("staffCompare.html", s1 = staff1, s2 = staff2, t1 = t1events, t2 = t2events)


@app.route('/update-data')
def upload_file():
    return render_template('update-data.html')


@app.route('/uploader', methods=['GET', 'POST'])
def update_data():
    if request.method == 'POST':
        f = request.files['file']
        if f.filename[-4:] == ".csv":
            filepath = "data/report.csv"

        elif f.filename[-4:] == ".xls":
            filepath = "data/report.xls"

        elif f.filename[-5:] == ".xlsx":
            filepath = "data/report.xlsx"
        else:
            raise FileTypeError("Filetype not recognised. Must use CSV, XLS or XLSX.")

        f.save(filepath)
        program.load_data(filepath)
        program.refresh_summary()
        return redirect('/')

program = EventsDB()
program.load_data(data_source)
program.refresh_summary()
app.run(debug=True, host='0.0.0.0', port=2000)
